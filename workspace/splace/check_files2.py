# -*- coding: utf-8 -*-
import sys
import io
import openpyxl
import msoffcrypto

out = open(r'C:\Users\남혜은\Desktop\파일구조.txt', 'w', encoding='utf-8')

# 1. 예약 파일
out.write("=" * 60 + "\n[1] 예약자관리 파일\n" + "=" * 60 + "\n")
try:
    with open(r'C:\Users\남혜은\Desktop\시간을담다베이커리_예약자관리_20260510_2207.xlsx', 'rb') as f:
        office_file = msoffcrypto.OfficeFile(f)
        if office_file.is_encrypted():
            office_file.load_key(password='kdsee1')
            dec = io.BytesIO()
            office_file.decrypt(dec)
            dec.seek(0)
            wb = openpyxl.load_workbook(dec)
        else:
            f.seek(0)
            wb = openpyxl.load_workbook(f)

    for sname in wb.sheetnames:
        ws = wb[sname]
        out.write(f"\n시트: {sname}  ({ws.max_row}행 x {ws.max_column}열)\n")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True)):
            out.write(f"  행{i+1}: {row}\n")
except Exception as e:
    out.write(f"오류: {e}\n")

# 2. 통계 파일
out.write("\n" + "=" * 60 + "\n[2] 스마트스토어 통계 파일\n" + "=" * 60 + "\n")
try:
    wb2 = openpyxl.load_workbook(r'C:\Users\남혜은\Desktop\스마트스토어 통계.xlsx', data_only=True)
    for sname in wb2.sheetnames:
        ws2 = wb2[sname]
        out.write(f"\n시트: {sname}  ({ws2.max_row}행 x {ws2.max_column}열)\n")
        for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=min(10, ws2.max_row), values_only=True)):
            out.write(f"  행{i+1}: {row}\n")
except Exception as e:
    out.write(f"오류: {e}\n")

out.write("\n완료\n")
out.close()
print("완료 - 파일구조.txt 확인하세요")
