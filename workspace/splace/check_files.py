# -*- coding: utf-8 -*-
import sys
import io

# 1. 예약 파일 읽기 (비밀번호: kdsee1)
print("=" * 60)
print("[1] 예약자관리 파일")
print("=" * 60)

try:
    import msoffcrypto
    import openpyxl

    with open(r'C:\Users\남혜은\Desktop\시간을담다베이커리_예약자관리_20260510_2207.xlsx', 'rb') as f:
        office_file = msoffcrypto.OfficeFile(f)
        if office_file.is_encrypted():
            office_file.load_key(password='kdsee1')
            decrypted = io.BytesIO()
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            wb = openpyxl.load_workbook(decrypted)
        else:
            f.seek(0)
            wb = openpyxl.load_workbook(f)

    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n시트: {sname}  ({ws.max_row}행 x {ws.max_column}열)")
        # 헤더 + 처음 5행
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            print(row)

except Exception as e:
    print(f"오류: {e}")

# 2. 통계 파일 읽기
print("\n" + "=" * 60)
print("[2] 스마트스토어 통계 파일")
print("=" * 60)

try:
    import openpyxl
    wb2 = openpyxl.load_workbook(r'C:\Users\남혜은\Desktop\스마트스토어 통계.xlsx')
    for sname in wb2.sheetnames:
        ws2 = wb2[sname]
        print(f"\n시트: {sname}  ({ws2.max_row}행 x {ws2.max_column}열)")
        for row in ws2.iter_rows(min_row=1, max_row=6, values_only=True):
            print(row)
except Exception as e:
    print(f"오류: {e}")

print("\n완료")
