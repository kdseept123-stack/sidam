# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import threading
import logging
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, 'config.json')
HISTORY_PATH = os.path.join(BASE_DIR, 'history.json')
LOG_PATH = os.path.join(BASE_DIR, 'automation.log')

logging.basicConfig(
    filename=LOG_PATH,
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s %(message)s',
    encoding='utf-8',
    force=True,
)

LOGEN_MAX_PRODUCT_LEN = 30

# Smart Store Excel column name candidates (row 2 is header row)
COLUMN_CANDIDATES = {
    '상품주문번호': ['상품주문번호'],
    '주문번호':     ['주문번호'],
    '수취인명':     ['수취인명', '수취인이름', '받는분성명'],
    '수취인전화번호': ['수취인연락처', '수취인전화번호', '전화번호'],
    '배송주소':     ['배송지', '수취인주소', '주소지', '받는분주소', '주소'],
    '우편번호':     ['우편번호', '배송지우편번호'],
    '상품명':       ['상품명'],
    '수량':         ['수량'],
    '배송메세지':   ['배송메세지', '배송메시지', '배송 메세지'],
    '구매자명':     ['구매자명'],
    '구매자전화번호': ['구매자전화번호', '구매자연락처', '구매자 전화번호'],
}


class LogenApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("로젠택배 자동화")
        self.root.geometry("1150x720")
        self.root.resizable(True, True)

        self.config = self._load_json(CONFIG_PATH, {
            'logen_id': '', 'logen_pw': '',
            'sender_name': '', 'sender_phone': '',
            'sender_zipcode': '', 'sender_address': '',
        })
        self.history = self._load_json(HISTORY_PATH, {})
        self.orders = []

        self._build_ui()

    # ──────────────────────────────────────────────────────────
    # Persistence helpers
    # ──────────────────────────────────────────────────────────

    def _load_json(self, path, default):
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return default

    def _save_json(self, path, data):
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    # ──────────────────────────────────────────────────────────
    # UI construction
    # ──────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Top bar ──────────────────────────────────────────
        top = tk.Frame(self.root, bd=1, relief='groove')
        top.pack(fill='x', padx=8, pady=(8, 2))

        tk.Button(top, text="📂 엑셀 파일 선택", command=self.load_excel,
                  font=('맑은 고딕', 11), padx=8, pady=4).pack(side='left', padx=6, pady=4)

        self.file_label = tk.Label(top, text="파일을 선택하세요",
                                   font=('맑은 고딕', 10), fg='gray')
        self.file_label.pack(side='left', padx=6)

        tk.Button(top, text="⚙ 설정", command=self.open_settings,
                  font=('맑은 고딕', 10), padx=6).pack(side='right', padx=6, pady=4)

        # ── Control bar ──────────────────────────────────────
        ctrl = tk.Frame(self.root)
        ctrl.pack(fill='x', padx=8, pady=2)

        self.all_var = tk.BooleanVar(value=True)
        tk.Checkbutton(ctrl, text="전체 선택 / 해제", variable=self.all_var,
                       command=self._toggle_all, font=('맑은 고딕', 10)).pack(side='left', padx=4)

        self.count_label = tk.Label(ctrl, text="", font=('맑은 고딕', 10))
        self.count_label.pack(side='left', padx=16)

        self.bundle_label = tk.Label(ctrl, text="", font=('맑은 고딕', 10), fg='#1565C0')
        self.bundle_label.pack(side='left')

        # ── Order table ──────────────────────────────────────
        tbl_frame = tk.Frame(self.root)
        tbl_frame.pack(fill='both', expand=True, padx=8, pady=4)

        cols = ('선택', '수취인명', '상품명', '수량', '구매자전화번호', '배송메세지', '상태')
        self.tree = ttk.Treeview(tbl_frame, columns=cols, show='headings', height=22)

        col_width = {
            '선택': 48, '수취인명': 90, '상품명': 320, '수량': 48,
            '구매자전화번호': 120, '배송메세지': 200, '상태': 80,
        }
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=col_width.get(c, 100), anchor='center' if c in ('선택', '수량') else 'w')

        vsb = ttk.Scrollbar(tbl_frame, orient='vertical', command=self.tree.yview)
        hsb = ttk.Scrollbar(tbl_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tbl_frame.columnconfigure(0, weight=1)
        tbl_frame.rowconfigure(0, weight=1)

        self.tree.bind('<Button-1>', self._on_row_click)
        self.tree.tag_configure('done',   background='#E8F5E9')
        self.tree.tag_configure('bundle', background='#E3F2FD')
        self.tree.tag_configure('fail',   background='#FFEBEE')

        # ── Action buttons ───────────────────────────────────
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(fill='x', padx=8, pady=4)

        self.start_btn = tk.Button(
            btn_frame, text="🚀 발송 시작", command=self.start_shipping,
            font=('맑은 고딕', 12, 'bold'), bg='#1976D2', fg='white',
            padx=14, pady=6, state='disabled')
        self.start_btn.pack(side='left', padx=5)

        self.save_btn = tk.Button(
            btn_frame, text="📊 스마트스토어 엑셀 저장", command=self.save_smartstore_excel,
            font=('맑은 고딕', 11), padx=10, pady=6, state='disabled')
        self.save_btn.pack(side='left', padx=5)

        self.retry_btn = tk.Button(
            btn_frame, text="🔄 실패 재시도", command=self.retry_failed,
            font=('맑은 고딕', 11), padx=10, pady=6, state='disabled')
        self.retry_btn.pack(side='left', padx=5)

        self.print_btn = tk.Button(
            btn_frame, text="🖨 목록 출력", command=self.print_order_list,
            font=('맑은 고딕', 11), padx=10, pady=6, state='disabled')
        self.print_btn.pack(side='left', padx=5)

        # ── Progress bar ─────────────────────────────────────
        self.progress = ttk.Progressbar(self.root, mode='determinate')
        self.progress.pack(fill='x', padx=8, pady=(2, 0))

        self.status_var = tk.StringVar(value="엑셀 파일을 선택하면 주문 목록이 표시됩니다.")
        tk.Label(self.root, textvariable=self.status_var,
                 font=('맑은 고딕', 10), anchor='w').pack(fill='x', padx=10, pady=(2, 6))

    # ──────────────────────────────────────────────────────────
    # Excel parsing
    # ──────────────────────────────────────────────────────────

    def load_excel(self):
        path = filedialog.askopenfilename(
            title="스마트스토어 주문 엑셀 선택",
            filetypes=[("Excel", "*.xlsx *.xls"), ("All", "*.*")])
        if not path:
            return

        # tkinter returns forward slashes; win32com needs backslashes
        path = os.path.normpath(path)

        self.file_label.config(text=os.path.basename(path), fg='black')
        self._set_status("엑셀 읽는 중...")
        self.root.update()

        try:
            self.orders = self._parse_excel(path)
            self._refresh_table()
            self.start_btn.config(state='normal')
            self.save_btn.config(state='normal')
            self.print_btn.config(state='normal')
            self._set_status(f"총 {len(self.orders)}건 로드 완료")
        except Exception as e:
            import traceback
            detail = traceback.format_exc()
            messagebox.showerror("엑셀 읽기 오류", f"{e}\n\n상세:\n{detail[-600:]}")
            self._set_status("엑셀 읽기 실패")

    def _parse_excel(self, filepath):
        import win32com.client as win32
        import subprocess

        # Kill any stale Excel processes to avoid COM conflicts
        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'],
                       capture_output=True)

        excel = win32.Dispatch('Excel.Application')
        excel.DisplayAlerts = False
        excel.Visible = False
        excel.Interactive = False
        wb = None
        try:
            wb = excel.Workbooks.Open(
                filepath,
                UpdateLinks=0,
                ReadOnly=True,
                IgnoreReadOnlyRecommended=True,
                Password='1111',
            )
            ws = wb.ActiveSheet
            rows = ws.UsedRange.Rows.Count
            cols = ws.UsedRange.Columns.Count

            if rows < 2 or cols < 1:
                raise ValueError("엑셀 파일에 데이터가 없습니다.")

            # Row 2 is the actual header row (row 1 is Naver's instruction text)
            headers = {}
            for c in range(1, cols + 1):
                v = ws.Cells(2, c).Value
                if v is not None:
                    headers[c] = str(v).strip()

            col_map = self._map_columns(headers)

            # 헤더 전체와 매핑 결과 로그
            logging.info("엑셀 헤더: %s", dict(headers))
            logging.info("컬럼 매핑: %s", {k: (v, headers.get(v)) for k, v in col_map.items() if v})

            detected = {k: v for k, v in col_map.items() if v}
            self._set_status(f"컬럼 인식: {list(detected.keys())}")

            orders = []
            for r in range(3, rows + 1):
                order = {field: '' for field in COLUMN_CANDIDATES}
                for field, col_idx in col_map.items():
                    if col_idx:
                        v = ws.Cells(r, col_idx).Value
                        if v is None:
                            order[field] = ''
                        elif field == '수량':
                            # 수량은 정수로 변환; 실패하면 원본 유지
                            try:
                                order[field] = str(int(float(str(v))))
                            except (ValueError, TypeError):
                                order[field] = str(v).strip()
                        else:
                            order[field] = str(v).strip()

                if not order['상품주문번호'] and not order['수취인명']:
                    continue  # empty row

                order['_selected'] = True
                order['_tracking'] = ''
                order['_failed'] = False
                orders.append(order)

            return orders
        finally:
            if wb is not None:
                try:
                    wb.Close(False)
                except Exception:
                    pass
            try:
                excel.Quit()
            except Exception:
                pass

    def _map_columns(self, headers):
        result = {field: None for field in COLUMN_CANDIDATES}
        for field, candidates in COLUMN_CANDIDATES.items():
            # Pass 1: exact match (prevents '수량' from matching '수량클레임 여부')
            for col_idx, col_name in headers.items():
                if col_name in candidates:
                    result[field] = col_idx
                    break
            if result[field]:
                continue
            # Pass 2: substring fallback
            for col_idx, col_name in headers.items():
                for cand in candidates:
                    if cand in col_name or col_name in cand:
                        result[field] = col_idx
                        break
                if result[field]:
                    break
        return result

    # ──────────────────────────────────────────────────────────
    # Table display
    # ──────────────────────────────────────────────────────────

    def _refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        bundle_map = self._build_bundle_map()
        bundle_count = sum(1 for v in bundle_map.values() if len(v) > 1)

        for i, o in enumerate(self.orders):
            key = self._bundle_key(o)
            is_bundle = len(bundle_map[key]) > 1
            is_done = o.get('상품주문번호', '') in self.history or bool(o['_tracking'])
            is_fail = o['_failed']

            tag = 'fail' if is_fail else ('done' if is_done else ('bundle' if is_bundle else ''))

            name = o.get('상품명', '')
            name_short = name[:40] + ('…' if len(name) > 40 else '')

            phone = o.get('구매자전화번호') or o.get('수취인전화번호', '')

            status = ''
            if is_fail:
                status = '❌ 실패'
            elif o['_tracking']:
                status = '✅ 완료'
            elif is_done:
                status = '✅ 기발송'
            elif is_bundle:
                status = '📦 묶음'

            self.tree.insert('', 'end', iid=str(i), tags=(tag,), values=(
                '☑' if o['_selected'] else '☐',
                o.get('수취인명', ''),
                name_short,
                o.get('수량', ''),
                phone,
                o.get('배송메세지', ''),
                status,
            ))

        selected = sum(1 for o in self.orders if o['_selected'])
        self.count_label.config(text=f"선택: {selected}건 / 전체: {len(self.orders)}건")
        self.bundle_label.config(
            text=f"  묶음배송 그룹: {bundle_count}개" if bundle_count else "")

    def _build_bundle_map(self):
        m = {}
        for o in self.orders:
            key = self._bundle_key(o)
            m.setdefault(key, []).append(o)
        return m

    def _bundle_key(self, order):
        return (order.get('수취인명', ''), order.get('수취인전화번호', ''))

    def _on_row_click(self, event):
        col = self.tree.identify_column(event.x)
        row = self.tree.identify_row(event.y)
        if col == '#1' and row:
            idx = int(row)
            self.orders[idx]['_selected'] = not self.orders[idx]['_selected']
            vals = list(self.tree.item(row, 'values'))
            vals[0] = '☑' if self.orders[idx]['_selected'] else '☐'
            self.tree.item(row, values=vals)
            selected = sum(1 for o in self.orders if o['_selected'])
            self.count_label.config(
                text=f"선택: {selected}건 / 전체: {len(self.orders)}건")

    def _toggle_all(self):
        state = self.all_var.get()
        for i, o in enumerate(self.orders):
            o['_selected'] = state
            vals = list(self.tree.item(str(i), 'values'))
            vals[0] = '☑' if state else '☐'
            self.tree.item(str(i), values=vals)
        selected = sum(1 for o in self.orders if o['_selected'])
        self.count_label.config(
            text=f"선택: {selected}건 / 전체: {len(self.orders)}건")

    # ──────────────────────────────────────────────────────────
    # Logen automation
    # ──────────────────────────────────────────────────────────

    def start_shipping(self):
        targets = [o for o in self.orders
                   if o['_selected'] and not o['_tracking']
                   and o.get('상품주문번호', '') not in self.history]
        if not targets:
            messagebox.showinfo("알림", "발송할 주문이 없습니다.\n(이미 처리됐거나 선택된 건이 없습니다.)")
            return

        if not self.config.get('logen_id') or not self.config.get('logen_pw'):
            messagebox.showwarning("설정 필요", "설정에서 로젠 로그인 정보를 먼저 입력해주세요.")
            self.open_settings()
            return

        self.start_btn.config(state='disabled')
        self.progress.config(maximum=len(targets), value=0)
        self._set_status(f"{len(targets)}건 발송 시작...")

        t = threading.Thread(target=self._automation_worker, args=(targets,), daemon=True)
        t.name = "logen-automation"
        t.start()

    def _automation_worker(self, targets):
        try:
            self._run_automation(targets)
        except Exception as e:
            import traceback
            detail = traceback.format_exc()
            logging.error("자동화 오류: %s\n%s", e, detail)
            self.root.after(0, lambda: messagebox.showerror(
                "자동화 오류",
                f"{e}\n\n상세 로그: {LOG_PATH}\n\n{detail[-600:]}"))
            self.root.after(0, lambda: self.start_btn.config(state='normal'))
            self._set_status(f"오류: {e}")

    def _run_automation(self, targets):
        import os, tempfile
        from playwright.sync_api import sync_playwright

        # 묶음 그룹 구성 (순서 유지)
        groups = {}
        for o in targets:
            key = self._bundle_key(o)
            groups.setdefault(key, []).append(o)
        group_list = list(groups.values())
        total = len(group_list)

        success = 0
        fail = 0

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False, slow_mo=300)
            context = browser.new_context()
            page = context.new_page()

            try:
                if not self._logen_login(page):
                    self.root.after(0, lambda: messagebox.showerror(
                        "로그인 실패",
                        "logis.ilogen.com 로그인에 실패했습니다.\n\n"
                        "⚙ 설정에서 아이디/비밀번호를 다시 확인해주세요."))
                    self.root.after(0, lambda: self.start_btn.config(state='normal'))
                    return

                self._set_status("복수건 폼 열기 중... (최대 60초)")
                logging.info("복수건 폼 열기 시도")
                frame = self._get_multi_order_frame(page)
                logging.info("복수건 폼 로드 완료: %s", frame.url)

                self._set_status(f"{total}건 엑셀 생성 중...")
                tmp_path = self._generate_bulk_excel(group_list)
                logging.info("임시 엑셀 생성: %s", tmp_path)

                try:
                    self._set_status(f"{total}건 업로드 및 전송 중...")
                    logging.info("업로드 및 서버전송 시작")
                    tracking_map = self._upload_and_submit_bulk(
                        frame, page, tmp_path, total)
                    logging.info("tracking_map: %s", tracking_map)
                finally:
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass

                # 결과를 orders에 반영
                for i, group in enumerate(group_list):
                    tracking_no = tracking_map.get(i)
                    rep = group[0]
                    name = rep.get('수취인명', '')
                    if tracking_no:
                        for o in group:
                            o['_tracking'] = tracking_no
                            o['_failed'] = False
                            self.history[o.get('상품주문번호', '')] = {
                                'tracking_no': tracking_no,
                                'recipient': name,
                                'date': datetime.now().isoformat(),
                            }
                        success += 1
                    else:
                        for o in group:
                            o['_failed'] = True
                        fail += 1

                self.root.after(0, lambda v=total: self.progress.config(value=v))

            finally:
                browser.close()

        self._save_json(HISTORY_PATH, self.history)
        self.root.after(0, self._refresh_table)
        self.root.after(0, lambda: self.start_btn.config(state='normal'))
        self.root.after(0, lambda: self.retry_btn.config(
            state='normal' if fail > 0 else 'disabled'))
        self.root.after(0, lambda: self.save_btn.config(state='normal'))

        msg = f"완료: {success}건 성공"
        if fail:
            msg += f", {fail}건 실패 → '실패 재시도' 버튼으로 재처리 가능"
        self._set_status(msg)
        self.root.after(0, lambda: messagebox.showinfo("발송 완료", msg))

    def _logen_login(self, page):
        """Login to logis.ilogen.com. Returns True on success."""
        page.goto('https://logis.ilogen.com/', wait_until='domcontentloaded')
        page.wait_for_timeout(2000)

        page.fill('[id="user.id"]', self.config['logen_id'])
        page.fill('[id="user.pw"]', self.config['logen_pw'])
        page.locator('[id="user.pw"]').press('Enter')
        page.wait_for_timeout(3000)

        # 팝업 닫기
        for close_text in ['닫기', '확인']:
            try:
                page.click(f'text={close_text}', timeout=2000)
                page.wait_for_timeout(500)
            except Exception:
                pass

        body = page.inner_text('body')
        if 'NoSuchUser' in body or '사용자 계정이 없습니다' in body:
            return False
        return True

    def _get_multi_order_frame(self, page):
        """복수건 메뉴를 열고 iframe 반환. 이미 열려 있으면 재사용."""
        # 이미 iframe이 있으면 바로 반환
        for f in page.frames:
            if 'lrm01f0040' in f.url:
                return f

        # 메뉴 검색창에 '복수건' 입력 (AJAX 검색 결과가 최상위 DOM에 렌더링됨)
        try:
            page.fill('#menuInput', '복수건', timeout=3000)
        except Exception:
            pass

        # 검색 결과 렌더링 대기 후 매초 클릭 시도 (최대 60초)
        # - 처음 2~3초는 AJAX 결과가 아직 없어서 클릭 무시됨
        # - 검색 결과가 뜨면 page.evaluate가 링크를 찾아 폼 열림
        JS_CLICK = """
            () => {
                const el = document.querySelector('a[title="주문등록/출력(복수건)"]');
                if (el) { el.click(); return true; }
                return false;
            }
        """
        for attempt in range(60):
            # 최상위 document (검색 결과 포함) 에서 클릭
            try:
                page.evaluate(JS_CLICK)
            except Exception:
                pass
            # 모든 frame 에서도 시도
            for f in page.frames:
                try:
                    f.evaluate(JS_CLICK)
                except Exception:
                    pass

            page.wait_for_timeout(1000)

            # 폼 로드 확인
            for f in page.frames:
                if 'lrm01f0040' in f.url:
                    self._set_status("복수건 폼 로드 완료")
                    f.wait_for_load_state('domcontentloaded')
                    page.wait_for_timeout(1000)
                    return f

            if attempt % 5 == 0:
                self._set_status(f"복수건 폼 대기 중... {attempt}초 경과")

        raise Exception("복수건 주문등록 폼 iframe을 찾을 수 없습니다. (60초 초과)\n메뉴에서 '주문등록/출력(복수건)' 권한이 있는지 확인하세요.")

    def _generate_bulk_excel(self, group_list):
        """A타입(제목없음) 형식의 임시 Excel 파일 생성. group_list 순서대로 행 추가."""
        import tempfile
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        for group in group_list:
            rep = group[0]

            # 주소 (배송주소 전체)
            addr = rep.get('배송주소', '')

            # 물품명 (묶음 시 합산)
            names = list({o.get('상품명', '') for o in group if o.get('상품명')})
            product_text = ', '.join(names)
            if len(product_text) > LOGEN_MAX_PRODUCT_LEN:
                product_text = product_text[:LOGEN_MAX_PRODUCT_LEN - 2] + '..'

            # 수량 합산
            try:
                total_qty = sum(int(o.get('수량', 1) or 1) for o in group)
            except Exception:
                total_qty = 1

            # A타입 컬럼(제목없음):
            # 수하인명, 우편번호, 수하인주소, 수하인전화번호, 수하인핸드폰번호,
            # 택배수량, 택배운임, 운임구분, 품목명, (공란), 배송메세지
            ws.append([
                rep.get('수취인명', ''),
                rep.get('우편번호', ''),
                addr,
                '',                              # 집전화 (공란)
                rep.get('수취인전화번호', ''),
                total_qty,
                '',                              # 택배운임 (공란 → 계약운임 자동)
                '010',                           # 운임구분: 선불
                product_text,
                '',                              # (공란)
                rep.get('배송메세지', '')[:40],
            ])

        tmp = tempfile.NamedTemporaryFile(
            suffix='.xlsx', delete=False, dir=tempfile.gettempdir())
        tmp.close()
        wb.save(tmp.name)
        return tmp.name

    def _upload_and_submit_bulk(self, frame, page, tmp_path, total):
        """복수건 폼: Excel 업로드 → 검증 → 서버전송 → 운송장번호 리스트 반환."""
        import re
        cfg = self.config

        # 접수일자 (오늘)
        today = datetime.now().strftime('%Y-%m-%d')
        try:
            frame.fill('#takeDate', today, timeout=3000)
        except Exception:
            pass

        # 송하인명
        try:
            frame.fill('#sndCustNm', cfg.get('sender_name', ''), timeout=3000)
        except Exception:
            pass

        # 송하인 휴대폰 3분할 (010-xxxx-xxxx)
        phone_raw = re.sub(r'\D', '', cfg.get('sender_phone', ''))
        if len(phone_raw) == 11:
            try:
                frame.fill('#cellNo1', phone_raw[:3])
                frame.fill('#cellNo2', phone_raw[3:7])
                frame.fill('#cellNo3', phone_raw[7:])
            except Exception:
                pass

        # Excel 타입: A타입 (value='100')
        try:
            frame.select_option('#cbExcelTy', value='100', timeout=3000)
            frame.wait_for_timeout(500)
        except Exception:
            pass

        # 합포장안함 (우리가 이미 묶음 처리함)
        try:
            frame.select_option('#cbMrgPackStandard', value='0', timeout=2000)
        except Exception:
            pass

        # ── 파일 업로드 ──────────────────────────────────────────
        uploaded = False
        logging.info("파일 업로드 시도: %s", tmp_path)
        try:
            with page.expect_file_chooser(timeout=6000) as fc_info:
                frame.click('text=파일열기', timeout=6000)
            fc_info.value.set_files(tmp_path)
            uploaded = True
            logging.info("파일선택창 업로드 성공")
        except Exception as e:
            logging.warning("파일선택창 업로드 실패: %s", e)

        if not uploaded:
            try:
                frame.locator('input[type="file"]').first.set_input_files(tmp_path)
                uploaded = True
                logging.info("hidden input 업로드 성공")
            except Exception as e:
                logging.error("hidden input 업로드 실패: %s", e)
                # 스크린샷 저장
                try:
                    ss_path = os.path.join(BASE_DIR, 'error_screenshot.png')
                    page.screenshot(path=ss_path)
                    logging.info("스크린샷 저장: %s", ss_path)
                except Exception:
                    pass
                raise Exception(f"파일 업로드 실패: {e}\n스크린샷: {BASE_DIR}\\error_screenshot.png")

        frame.wait_for_timeout(2000)

        # ── 엑셀검증 ─────────────────────────────────────────────
        logging.info("엑셀검증 클릭")
        frame.click('text=엑셀검증', timeout=5000)
        frame.wait_for_timeout(3000)

        body = frame.inner_text('body')
        logging.debug("검증 후 body(500자): %s", body[:500])
        chk_m = re.search(r'정상-(\d+)건, 오류-(\d+)건', body)
        ok_cnt  = int(chk_m.group(1)) if chk_m else 0
        err_cnt = int(chk_m.group(2)) if chk_m else 0
        self._set_status(f"검증 결과: 정상 {ok_cnt}건, 오류 {err_cnt}건")
        logging.info("검증 결과: 정상=%d, 오류=%d", ok_cnt, err_cnt)

        if ok_cnt == 0:
            try:
                ss_path = os.path.join(BASE_DIR, 'error_screenshot.png')
                page.screenshot(path=ss_path)
            except Exception:
                pass
            raise Exception(f"엑셀 검증 실패: 정상 0건 / 오류 {err_cnt}건\n스크린샷: {BASE_DIR}\\error_screenshot.png")

        # ── 2.서버전송 ────────────────────────────────────────────
        # 브라우저 다이얼로그(alert/confirm)를 자동 수락 — page.click('text=확인')은
        # 페이지 내 다른 버튼을 눌러 이중 전송을 일으킬 수 있으므로 사용 금지
        def _auto_accept(dialog):
            logging.info("다이얼로그 자동수락: %s", dialog.message)
            dialog.accept()

        page.on('dialog', _auto_accept)

        logging.info("서버전송 클릭")
        frame.click('text=서버전송', timeout=5000)

        # 전송 처리 대기 (dialog handler는 운송장출력까지 유지)
        frame.wait_for_timeout(8000)

        # ── 변환완료 탭으로 전환 ─────────────────────────────────
        try:
            frame.click('text=변환완료', timeout=3000)
            frame.wait_for_timeout(2000)
        except Exception:
            pass

        # ── 모든 행 체크박스 선택 ────────────────────────────────
        try:
            frame.evaluate("""
                () => {
                    document.querySelectorAll('input[type="checkbox"]').forEach(cb => {
                        if (!cb.checked) cb.click();
                    });
                }
            """)
            frame.wait_for_timeout(500)
        except Exception:
            pass

        # ── 운송장출력 클릭 (dialog handler 활성 상태에서) ───────
        # 다이얼로그(확인/예) 자동 수락 후 인쇄 진행
        logging.info("운송장출력 클릭 시도")
        try:
            clicked = frame.evaluate("""
                () => {
                    const els = Array.from(document.querySelectorAll('a, button, input[type="button"], span'));
                    const el = els.find(e => (e.innerText || e.value || e.textContent || '').includes('운송장출력'));
                    if (el) { el.click(); return true; }
                    return false;
                }
            """)
            logging.info("운송장출력 JS 클릭: %s", clicked)
        except Exception as e:
            logging.warning("운송장출력 JS 클릭 실패: %s", e)
            try:
                frame.click('text=운송장출력', timeout=5000)
            except Exception:
                pass

        # 인쇄 처리 및 그리드 갱신 대기
        frame.wait_for_timeout(12000)
        page.remove_listener('dialog', _auto_accept)

        # ── 엑셀저장으로 운송장번호 추출 (가장 신뢰성 높음) ──────
        import tempfile as _tf
        tracking_nos = []
        dl_path = None
        try:
            dl_path = os.path.join(_tf.gettempdir(),
                                   f'logen_result_{datetime.now():%Y%m%d_%H%M%S}.xlsx')
            with page.expect_download(timeout=15000) as dl_info:
                try:
                    frame.click('text=엑셀저장', timeout=5000)
                except Exception:
                    page.click('text=엑셀저장', timeout=3000)
            dl_info.value.save_as(dl_path)
            logging.info("엑셀저장 다운로드: %s", dl_path)

            import openpyxl as _opxl
            wb2 = _opxl.load_workbook(dl_path)
            for ws2 in wb2.worksheets:
                for row in ws2.iter_rows():
                    for cell in row:
                        v = str(cell.value or '').strip()
                        if re.match(r'^\d{10,13}$', v) and not re.match(r'^01[016789]\d{7,8}$', v):
                            if v not in tracking_nos:
                                tracking_nos.append(v)
            logging.info("엑셀저장에서 운송장번호 후보: %s", tracking_nos)
        except Exception as e:
            logging.warning("엑셀저장 실패: %s", e)
        finally:
            if dl_path:
                try:
                    os.remove(dl_path)
                except Exception:
                    pass

        # fallback: 변환완료 탭 그리드에서 JS로 읽기
        if not tracking_nos:
            tracking_nos = frame.evaluate("""
                () => {
                    const seen = new Set();
                    const results = [];
                    const add = v => {
                        if (v && !seen.has(v) && !/^01[016789]/.test(v)) {
                            seen.add(v); results.push(v);
                        }
                    };
                    (document.body.innerText.match(/\\b\\d{10,13}\\b/g) || []).forEach(add);
                    document.querySelectorAll('td, input').forEach(el => {
                        const v = (el.innerText || el.value || '').trim();
                        if (/^\\d{10,13}$/.test(v)) add(v);
                    });
                    return results;
                }
            """)
            logging.info("그리드 fallback 운송장번호 후보: %s", tracking_nos)

        # 중복 제거 (순서 유지)
        seen = set()
        unique_nos = []
        for t in tracking_nos:
            if t not in seen:
                seen.add(t)
                unique_nos.append(t)

        # index → tracking_no 매핑
        tracking_map = {i: unique_nos[i] for i in range(min(total, len(unique_nos)))}
        self._set_status(f"운송장번호 {len(unique_nos)}건 확인됨")
        logging.info("최종 tracking_map: %s", tracking_map)

        return tracking_map

    def retry_failed(self):
        for o in self.orders:
            if o['_failed']:
                o['_selected'] = True
                o['_failed'] = False
        self._refresh_table()
        self.start_shipping()

    # ──────────────────────────────────────────────────────────
    # Smart Store upload Excel
    # ──────────────────────────────────────────────────────────

    def print_order_list(self):
        targets = [o for o in self.orders if o['_selected']]
        if not targets:
            messagebox.showinfo("알림", "선택된 주문이 없습니다.")
            return

        bundle_map = self._build_bundle_map()
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        total_qty = sum(int(o.get('수량', 1) or 1) for o in targets)

        rows_html = ''
        for i, o in enumerate(targets, 1):
            key = self._bundle_key(o)
            is_bundle = len(bundle_map[key]) > 1
            status = o.get('_tracking') or ('묶음' if is_bundle else '')
            bg = ' style="background:#E3F2FD"' if is_bundle else ''
            rows_html += (
                f'<tr{bg}>'
                f'<td>{i}</td>'
                f'<td>{o.get("수취인명","")}</td>'
                f'<td class="left">{o.get("상품명","")[:50]}</td>'
                f'<td>{o.get("수량","")}</td>'
                f'<td>{o.get("수취인전화번호","")}</td>'
                f'<td class="left">{o.get("배송메세지","")[:30]}</td>'
                f'</tr>\n'
            )

        html = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<title>발송 목록</title>
<style>
  body {{ font-family: '맑은 고딕', sans-serif; font-size: 11px; margin: 16px; }}
  h2 {{ font-size: 14px; margin-bottom: 4px; }}
  p  {{ font-size: 11px; color: #555; margin: 2px 0 8px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ border: 1px solid #aaa; padding: 3px 5px; text-align: center; white-space: nowrap; }}
  th {{ background: #1976D2; color: white; font-size: 11px; }}
  td.left {{ text-align: left; }}
  tr:nth-child(even) {{ background: #f5f5f5; }}
  @media print {{ body {{ margin: 4px; }} }}
</style>
</head><body>
<h2>발송 목록 — {now}</h2>
<p>선택 {len(targets)}건 · 총 수량 {total_qty}개</p>
<table>
<thead><tr>
  <th>No.</th><th>수취인</th><th>상품명</th><th>수량</th>
  <th>전화번호</th><th>배송메세지</th>
</tr></thead>
<tbody>
{rows_html}</tbody>
</table>
<script>window.onload = function(){{ window.print(); }};</script>
</body></html>"""

        import tempfile, webbrowser
        tmp = tempfile.NamedTemporaryFile(
            suffix='.html', delete=False, mode='w', encoding='utf-8')
        tmp.write(html)
        tmp.close()
        webbrowser.open(f'file:///{tmp.name.replace(chr(92), "/")}')

    def save_smartstore_excel(self):
        done = [o for o in self.orders if o['_tracking']]
        if not done:
            messagebox.showinfo("알림", "완료된 발송 건이 없습니다.")
            return

        path = filedialog.asksaveasfilename(
            title="스마트스토어 송장 업로드 엑셀 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"송장등록_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        if not path:
            return

        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "송장번호"
            ws.append(['상품주문번호', '배송방법', '택배사', '송장번호'])

            for o in done:
                ws.append([
                    o.get('상품주문번호', ''),
                    '택배',
                    '로젠택배',
                    o['_tracking'],
                ])

            wb.save(path)
            messagebox.showinfo("저장 완료", f"저장 완료:\n{path}")
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))

    # ──────────────────────────────────────────────────────────
    # Settings dialog
    # ──────────────────────────────────────────────────────────

    def open_settings(self):
        win = tk.Toplevel(self.root)
        win.title("설정")
        win.geometry("440x420")
        win.resizable(False, False)
        win.grab_set()

        def section(text, row):
            tk.Label(win, text=text, font=('맑은 고딕', 11, 'bold'),
                     fg='#1565C0').grid(row=row, column=0, columnspan=2,
                                        sticky='w', padx=16, pady=(12, 4))

        def field(label, key, row, show=''):
            tk.Label(win, text=label, font=('맑은 고딕', 10)).grid(
                row=row, column=0, sticky='e', padx=(16, 6), pady=4)
            e = tk.Entry(win, width=28, show=show)
            e.insert(0, self.config.get(key, ''))
            e.grid(row=row, column=1, sticky='w', pady=4)
            return e

        section("로젠택배 로그인", 0)
        entries = {
            'logen_id': field("아이디", 'logen_id', 1),
            'logen_pw': field("비밀번호", 'logen_pw', 2, show='*'),
        }

        section("발송인 정보", 3)
        for i, (label, key) in enumerate([
            ("이름", 'sender_name'),
            ("전화번호", 'sender_phone'),
            ("우편번호", 'sender_zipcode'),
            ("주소", 'sender_address'),
        ], 4):
            entries[key] = field(label, key, i)

        def save():
            for k, e in entries.items():
                self.config[k] = e.get().strip()
            self._save_json(CONFIG_PATH, self.config)
            win.destroy()
            messagebox.showinfo("저장", "설정이 저장되었습니다.")

        tk.Button(win, text="저장", command=save,
                  font=('맑은 고딕', 11), bg='#1976D2', fg='white',
                  width=12, pady=4).grid(row=9, column=0, columnspan=2, pady=16)

    # ──────────────────────────────────────────────────────────
    # Helpers
    # ──────────────────────────────────────────────────────────

    def _set_status(self, msg):
        self.root.after(0, lambda: self.status_var.set(msg))

    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    app = LogenApp()
    app.run()
