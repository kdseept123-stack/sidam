# -*- coding: utf-8 -*-
"""
시간을담다베이커리 - 스마트플레이스 예약 관리
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, Alignment
import msoffcrypto
import io, os, glob, tempfile, datetime, webbrowser, subprocess
from collections import defaultdict, OrderedDict

DESKTOP      = r'C:\Users\남혜은\Desktop'

def open_in_browser(path):
    url = 'file:///' + path.replace(os.sep, '/')
    browsers = [
        r'C:\Program Files\Google\Chrome\Application\chrome.exe',
        r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe',
        r'C:\Program Files\Microsoft\Edge\Application\msedge.exe',
        r'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe',
    ]
    for b in browsers:
        if os.path.exists(b):
            subprocess.Popen([b, url])
            return
    webbrowser.open(url)
STATS_FILE   = os.path.join(DESKTOP, '스마트스토어 통계.xlsx')
PASSWORD     = 'kdsee1'
STORE_NAME   = '시간을담다베이커리'
HYUNGMAN_SHEET  = '형만님'
STATS_HEADER_ROW = 13   # ' 스마트플레이스픽업' 헤더 행
STATS_START_ROW  = 14   # 제품 데이터 시작 행

# 예약 파일 열 인덱스 (0-based, 데이터는 4행부터)
C_BOOK_NO  = 0   # 예약번호
C_ORDER_NO = 3   # 주문번호
C_STATUS   = 5   # 상태
C_RESERVER = 7   # 예약자
C_PHONE    = 8   # 전화번호
C_USE_DATE = 13  # 이용일시
C_OPTION   = 17  # 가격분류 및 옵션 (상품명)
C_PRICE    = 18  # 결제금액 (항목별)
C_FINAL    = 20  # 실결제금액 (주문 전체)
C_PAYMENT  = 21  # 결제수단
C_REQUEST  = 27  # 예약자입력정보1 (컷팅/노컷)

# ── 파일 처리 ─────────────────────────────────────────────

def find_reservation_file():
    files = glob.glob(os.path.join(DESKTOP, f'{STORE_NAME}_예약자관리_*.xlsx'))
    return max(files, key=os.path.getmtime) if files else None

def read_reservation(path):
    with open(path, 'rb') as f:
        of = msoffcrypto.OfficeFile(f)
        if of.is_encrypted():
            of.load_key(password=PASSWORD)
            dec = io.BytesIO()
            of.decrypt(dec)
            dec.seek(0)
            wb = openpyxl.load_workbook(dec)
        else:
            f.seek(0)
            wb = openpyxl.load_workbook(f)
    ws = wb['Report']
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[C_ORDER_NO]:
            rows.append(row)
    return rows

def group_orders(rows):
    orders = OrderedDict()
    for row in rows:
        key = row[C_ORDER_NO]
        if key not in orders:
            orders[key] = dict(
                order_no   = key,
                status     = row[C_STATUS] or '',
                reserver   = row[C_RESERVER] or '',
                phone      = row[C_PHONE] or '',
                use_date   = str(row[C_USE_DATE] or ''),
                final      = row[C_FINAL] or 0,
                payment    = row[C_PAYMENT] or '',
                request    = row[C_REQUEST] or '',
                items      = [],
            )
        opt   = row[C_OPTION]
        price = row[C_PRICE] or 0
        if opt:
            orders[key]['items'].append({'option': opt, 'price': price})
    return list(orders.values())

def calc_stats(orders):
    counts = defaultdict(int)
    for o in orders:
        if '취소' in str(o['status']):
            continue
        for item in o['items']:
            counts[item['option']] += 1
    return dict(counts)

# ── 통계 파일 업데이트 ────────────────────────────────────

def update_stats_file(counts, date_label):
    wb = openpyxl.load_workbook(STATS_FILE)

    # 기존 플레이스예약 시트 있으면 제거
    if '플레이스예약' in wb.sheetnames:
        del wb['플레이스예약']

    ws = wb[HYUNGMAN_SHEET]

    # 헤더 행에 업데이트 날짜 기록
    ws.cell(STATS_HEADER_ROW, 1).value = f' 스마트플레이스픽업  ({date_label})'

    # A열 제품명 읽어서 B열 수량 업데이트 (정확히 일치하는 항목만)
    total = 0
    matched_names = set()
    for row_num in range(STATS_START_ROW, ws.max_row + 1):
        a_val = ws.cell(row_num, 1).value
        if a_val is None:
            break  # 제품 목록 끝
        sheet_name = str(a_val).strip()
        if not sheet_name:
            break

        matched = counts.get(sheet_name, 0)
        if matched:
            matched_names.add(sheet_name)
        ws.cell(row_num, 2).value = matched
        total += matched

    # 통계 파일에 없는 예약 제품 (누락 항목)
    unmatched = {k: v for k, v in counts.items() if k not in matched_names}

    wb.save(STATS_FILE)
    return total, unmatched

# ── 전체 예약 목록 출력 ───────────────────────────────────

def make_order_list_html(orders):
    today = datetime.date.today().strftime('%Y년 %m월 %d일')
    active = [o for o in orders if '취소' not in str(o['status'])]
    rows_html = ''
    for i, o in enumerate(orders):
        is_cancel = '취소' in str(o['status'])
        grouped = {}
        for item in o['items']:
            grouped[item['option']] = grouped.get(item['option'], 0) + 1
        items_str = ' / '.join(f'{k} {v}개' for k, v in grouped.items()) or '-'
        row_class = 'cancel' if is_cancel else ('odd' if i % 2 else 'even')
        status_tag = f' <span class="tag">[{o["status"]}]</span>' if is_cancel else ''
        req = o['request'] or ''
        rows_html += f'''
        <tr class="{row_class}">
          <td class="center">{i + 1}</td>
          <td>{o["reserver"]}<br><small>{o["phone"]}</small></td>
          <td class="center">{o["use_date"]}</td>
          <td>{items_str}</td>
          <td class="right">{o["final"]:,}원</td>
          <td class="center">{o["payment"]}</td>
          <td>{req}{status_tag}</td>
        </tr>'''

    return f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>{STORE_NAME} 예약 목록</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family: "맑은 고딕", "Malgun Gothic", sans-serif; padding:20px; font-size:12px; color:#222; }}
  h2 {{ text-align:center; font-size:17px; margin-bottom:4px; }}
  .subtitle {{ text-align:center; color:#666; margin-bottom:16px; font-size:11px; }}
  table {{ width:100%; border-collapse:collapse; }}
  th {{ background:#333; color:#fff; padding:7px 8px; font-size:11px; border:1px solid #222; }}
  td {{ padding:5px 8px; border:1px solid #ccc; vertical-align:middle; font-size:11px; }}
  .center {{ text-align:center; }}
  .right {{ text-align:right; }}
  .odd {{ background:#f9f9f9; }}
  .even {{ background:#ffffff; }}
  .cancel {{ color:#bbb; background:#f5f5f5; }}
  .cancel td {{ color:#bbb; }}
  .tag {{ font-size:10px; color:#e55; }}
  small {{ color:#999; font-size:10px; }}
  @media print {{
    body {{ padding:4mm; }}
    @page {{ margin:10mm; size:A4; }}
    .cancel {{ display:none; }}
  }}
</style>
</head>
<body>
<h2>🥖 {STORE_NAME} &nbsp; 예약 목록</h2>
<div class="subtitle">출력일: {today} &nbsp;|&nbsp; 전체 {len(orders)}건 (유효 {len(active)}건 / 취소 {len(orders) - len(active)}건)</div>
<table>
  <tr>
    <th style="width:36px">번호</th>
    <th style="width:130px">예약자 / 연락처</th>
    <th style="width:150px">픽업일시</th>
    <th>상품</th>
    <th style="width:90px">결제금액</th>
    <th style="width:80px">결제수단</th>
    <th style="width:130px">요청 / 상태</th>
  </tr>
  {rows_html}
</table>
<script>window.onload = function() {{ window.print(); }}</script>
</body></html>'''


# ── 영수증 ────────────────────────────────────────────────

def make_receipt_html(order):
    items_html = ''
    grouped = defaultdict(lambda: {'price': 0, 'count': 0})
    for item in order['items']:
        grouped[item['option']]['price'] = item['price']
        grouped[item['option']]['count'] += 1

    for opt, v in grouped.items():
        subtotal = v['price'] * v['count']
        items_html += f'''
        <tr>
          <td>{opt}</td>
          <td class="num">{v["count"]}</td>
          <td class="num">{v["price"]:,}원</td>
          <td class="num">{subtotal:,}원</td>
        </tr>'''

    req = f'<p class="req">요청: {order["request"]}</p>' if order['request'] else ''

    return f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family: "맑은 고딕", sans-serif; width:300px; margin:0 auto; padding:16px; font-size:13px; }}
  h2 {{ text-align:center; font-size:15px; border-bottom:2px solid #333; padding-bottom:6px; margin-bottom:10px; }}
  .info {{ margin-bottom:10px; line-height:1.7; }}
  table {{ width:100%; border-collapse:collapse; margin:10px 0; }}
  th {{ background:#333; color:#fff; padding:4px 6px; font-size:11px; }}
  td {{ padding:4px 6px; border-bottom:1px solid #eee; font-size:12px; }}
  .num {{ text-align:right; }}
  .total {{ font-weight:bold; font-size:14px; text-align:right; border-top:2px solid #333; padding-top:6px; margin-top:6px; }}
  .req {{ margin-top:8px; color:#555; font-size:12px; }}
  .footer {{ text-align:center; margin-top:14px; font-size:12px; color:#777; border-top:1px solid #ccc; padding-top:8px; }}
</style>
</head><body>
<h2>🥖 {STORE_NAME}</h2>
<div class="info">
  <b>예약자:</b> {order["reserver"]}<br>
  <b>연락처:</b> {order["phone"]}<br>
  <b>픽업일시:</b> {order["use_date"]}
</div>
<table>
  <tr><th>상품</th><th>수량</th><th>단가</th><th>소계</th></tr>
  {items_html}
</table>
<div class="total">실결제금액: {order["final"]:,}원</div>
<div class="info" style="margin-top:6px">결제수단: {order["payment"]}</div>
{req}
<div class="footer">감사합니다 ♡</div>
</body></html>'''

def print_receipt(order):
    html = make_receipt_html(order)
    tmp = tempfile.NamedTemporaryFile(
        mode='w', suffix='.html', encoding='utf-8', delete=False,
        dir=tempfile.gettempdir()
    )
    tmp.write(html)
    tmp.close()
    open_in_browser(tmp.name)

# ── GUI ──────────────────────────────────────────────────

BG_DARK   = '#1a1a2e'
BG_CARD   = '#0f0f1a'
BG_ITEM   = '#2a2a4e'
ORANGE    = '#ff6b35'
WHITE     = '#ffffff'
GRAY      = '#aaaacc'
FONT_BODY = ('맑은 고딕', 10)
FONT_HEAD = ('맑은 고딕', 11, 'bold')

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f'{STORE_NAME}  예약 관리')
        self.geometry('960x680')
        self.configure(bg=BG_DARK)
        self.orders = []
        self.stats  = {}
        self._setup_style()
        self._build_ui()
        self._load()

    def _setup_style(self):
        s = ttk.Style()
        s.theme_use('clam')
        s.configure('TNotebook', background=BG_DARK, borderwidth=0)
        s.configure('TNotebook.Tab', background=BG_ITEM, foreground=WHITE,
                    font=FONT_BODY, padding=[14, 6])
        s.map('TNotebook.Tab', background=[('selected', ORANGE)])
        s.configure('Tree.Treeview', background=BG_CARD, foreground=WHITE,
                    fieldbackground=BG_CARD, font=FONT_BODY, rowheight=38)
        s.configure('Tree.Treeview.Heading', background=BG_ITEM,
                    foreground=WHITE, font=FONT_HEAD)
        s.map('Tree.Treeview', background=[('selected', '#3a3a6e')])

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg=ORANGE, padx=16, pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text=f'🥖  {STORE_NAME}  예약 관리',
                 font=('맑은 고딕', 15, 'bold'), fg=WHITE, bg=ORANGE).pack(side='left')

        # Toolbar
        bar = tk.Frame(self, bg=BG_DARK, padx=12, pady=8)
        bar.pack(fill='x')
        for txt, cmd in [('🔄  새로고침 (최신파일)', self._load),
                          ('📂  파일 직접 선택', self._load_pick),
                          ('📊  통계 파일 업데이트', self._update_stats),
                          ('🖨  예약 목록 인쇄', self._print_all_orders)]:
            tk.Button(bar, text=txt, command=cmd,
                      bg=BG_ITEM, fg=WHITE, activebackground=ORANGE,
                      activeforeground=WHITE, relief='flat',
                      font=FONT_BODY, padx=12, pady=5, cursor='hand2'
                      ).pack(side='left', padx=4)
        self.lbl_status = tk.Label(bar, text='', fg=GRAY, bg=BG_DARK, font=('맑은 고딕', 9))
        self.lbl_status.pack(side='left', padx=12)

        # Notebook
        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=8, pady=6)

        frm1 = tk.Frame(nb, bg=BG_DARK); nb.add(frm1, text='  예약 목록  ')
        frm2 = tk.Frame(nb, bg=BG_DARK); nb.add(frm2, text='  제품 통계  ')
        frm3 = tk.Frame(nb, bg=BG_DARK); nb.add(frm3, text='  프린터 설정  ')

        self._build_order_tab(frm1)
        self._build_stat_tab(frm2)
        self._build_printer_tab(frm3)

    # ── 예약 목록 탭 ─────────────────────────────────────

    def _build_order_tab(self, parent):
        cols = ('예약자 / 전화번호', '픽업일시', '상품 목록', '실결제금액', '요청/컷팅', '영수증')
        widths = [150, 170, 260, 110, 130, 80]

        frame = tk.Frame(parent, bg=BG_DARK)
        frame.pack(fill='both', expand=True, padx=6, pady=6)

        self.tree = ttk.Treeview(frame, columns=cols, show='headings',
                                  style='Tree.Treeview', selectmode='browse')
        for col, w in zip(cols, widths):
            anchor = 'w' if col in ('상품 목록', '요청/컷팅', '예약자 / 전화번호') else 'center'
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor=anchor, stretch=col=='상품 목록')

        sb = ttk.Scrollbar(frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        self.tree.tag_configure('odd',    background='#131325')
        self.tree.tag_configure('even',   background=BG_CARD)
        self.tree.tag_configure('cancel', background='#2e1a1a', foreground='#888888')

        self.tree.bind('<ButtonRelease-1>', self._on_tree_click)

    # ── 프린터 설정 탭 ───────────────────────────────────

    def _build_printer_tab(self, parent):
        canvas = tk.Canvas(parent, bg=BG_DARK, highlightthickness=0)
        sb = ttk.Scrollbar(parent, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)

        inner = tk.Frame(canvas, bg=BG_DARK)
        win_id = canvas.create_window((0, 0), window=inner, anchor='nw')
        inner.bind('<Configure>', lambda e: canvas.configure(
            scrollregion=canvas.bbox('all')))
        canvas.bind('<Configure>', lambda e: canvas.itemconfig(win_id, width=e.width))

        def section(title):
            tk.Label(inner, text=title, font=('맑은 고딕', 12, 'bold'),
                     fg=ORANGE, bg=BG_DARK, anchor='w').pack(fill='x', padx=20, pady=(18, 4))
            tk.Frame(inner, bg=ORANGE, height=2).pack(fill='x', padx=20)

        def body(text):
            tk.Label(inner, text=text, font=('맑은 고딕', 10),
                     fg=WHITE, bg=BG_CARD, anchor='w', justify='left',
                     wraplength=700, padx=16, pady=10).pack(fill='x', padx=20, pady=4)

        def step(num, text):
            f = tk.Frame(inner, bg=BG_DARK)
            f.pack(fill='x', padx=20, pady=3)
            tk.Label(f, text=f' {num} ', font=('맑은 고딕', 10, 'bold'),
                     fg=WHITE, bg=ORANGE, width=3).pack(side='left')
            tk.Label(f, text=text, font=('맑은 고딕', 10),
                     fg=WHITE, bg=BG_DARK, anchor='w', justify='left',
                     wraplength=680).pack(side='left', padx=10, fill='x', expand=True)

        # ── 영수증 출력 방법 ──
        section('🖨  영수증 출력 방법')
        body('목록에서 🖨 출력 버튼을 클릭하면 미리보기 창이 열립니다.\n'
             '미리보기에서 내용 확인 후 [인쇄하기] 버튼을 누르면 브라우저가 열리고,\n'
             'Ctrl + P 를 눌러 프린터를 선택해 인쇄하세요.')

        # ── USB 연결 ──
        section('🔌  키오스크 프린터가 USB로 연결된 경우')
        step(1, '키오스크와 이 컴퓨터를 USB 케이블로 연결하거나,\n'
                '영수증 프린터를 직접 이 컴퓨터의 USB 포트에 꽂으세요.')
        step(2, '윈도우 시작 → 설정 → 블루투스 및 디바이스 → 프린터 및 스캐너 를 엽니다.')
        step(3, '영수증 프린터 이름을 클릭 → [기본값으로 설정] 을 누릅니다.')
        step(4, '이 앱에서 영수증 출력 후 브라우저의 인쇄 화면에서\n'
                '프린터 목록에 해당 프린터가 보이면 선택해 인쇄하세요.')

        # ── 네트워크(Wi-Fi/LAN) 연결 ──
        section('📡  키오스크 프린터가 같은 Wi-Fi / LAN에 연결된 경우')
        step(1, '윈도우 시작 → 설정 → 블루투스 및 디바이스 → 프린터 및 스캐너 로 이동합니다.')
        step(2, '[프린터 또는 스캐너 추가] 버튼을 클릭합니다.')
        step(3, '목록에 프린터가 나타나면 클릭해서 추가합니다.\n'
                '목록에 없으면 [원하는 프린터가 목록에 없습니다] 클릭 →\n'
                '"IP 주소로 프린터 추가" 선택 → 키오스크 프린터 IP를 입력합니다.')
        step(4, '추가된 프린터를 [기본값으로 설정] 합니다.')

        # ── 프린터 IP 확인 방법 ──
        section('🔍  키오스크 프린터 IP 주소 확인 방법')
        step(1, '키오스크 프린터에서 자체 테스트 인쇄를 합니다.\n'
                '(보통 프린터 뚜껑 안 버튼을 길게 누르면 테스트 용지가 출력됩니다)')
        step(2, '출력된 용지에서 IP Address 항목을 확인합니다.\n'
                '예: 192.168.0.50 형태로 나와 있습니다.')
        step(3, '그 IP를 위의 "IP 주소로 프린터 추가" 단계에서 입력하면 됩니다.')

        # ── 연결 모르는 경우 ──
        section('❓  연결 방법을 모르는 경우')
        body('키오스크 옆면이나 뒷면에 붙어 있는 모델명 스티커를 확인하세요.\n'
             '모델명(예: Bixolon SRP-350, Epson TM-T88)을 알려주시면\n'
             '더 정확한 연결 방법을 안내해드릴 수 있습니다.')

        tk.Label(inner, text='', bg=BG_DARK).pack(pady=16)

    # ── 통계 탭 ──────────────────────────────────────────

    def _build_stat_tab(self, parent):
        self.txt_stat = tk.Text(parent, bg=BG_CARD, fg=WHITE,
                                 font=('맑은 고딕', 12), padx=20, pady=16,
                                 relief='flat', state='disabled', wrap='none')
        sb = ttk.Scrollbar(parent, orient='vertical', command=self.txt_stat.yview)
        self.txt_stat.configure(yscrollcommand=sb.set)
        self.txt_stat.pack(side='left', fill='both', expand=True, padx=6, pady=6)
        sb.pack(side='right', fill='y', pady=6)

        self.txt_stat.tag_configure('header', font=('맑은 고딕', 13, 'bold'), foreground=ORANGE)
        self.txt_stat.tag_configure('item',   font=('맑은 고딕', 12))
        self.txt_stat.tag_configure('total',  font=('맑은 고딕', 12, 'bold'), foreground=ORANGE)
        self.txt_stat.tag_configure('dim',    foreground=GRAY)

    # ── 데이터 로드 ───────────────────────────────────────

    def _load_pick(self):
        path = filedialog.askopenfilename(
            title='예약 파일 선택',
            initialdir=DESKTOP,
            filetypes=[('Excel 파일', '*.xlsx'), ('모든 파일', '*.*')]
        )
        if path:
            self._load_from(path)

    def _load(self):
        path = find_reservation_file()
        if not path:
            messagebox.showerror('오류', f'데스크탑에서\n{STORE_NAME}_예약자관리_*.xlsx 파일을 찾을 수 없습니다.')
            return
        self._load_from(path)

    def _load_from(self, path):
        try:
            self.lbl_status.config(text='파일 읽는 중...')
            self.update()

            rows = read_reservation(path)
            self.orders = group_orders(rows)
            self.stats  = calc_stats(self.orders)
            fname = os.path.basename(path)

            self._refresh_tree()
            self._refresh_stat()

            # 통계 파일 자동 저장
            self._auto_save_stats(fname)

        except Exception as e:
            messagebox.showerror('파일 오류', str(e))

    def _auto_save_stats(self, fname):
        if not self.stats:
            self.lbl_status.config(text=f'{fname}  |  총 {len(self.orders)}건  |  ⚠ 예약 데이터 없음')
            return
        if not os.path.exists(STATS_FILE):
            self.lbl_status.config(text=f'{fname}  |  총 {len(self.orders)}건  |  ⚠ 통계 파일 없음')
            return
        try:
            today = datetime.date.today().strftime('%Y년 %m월 %d일')
            app_total = sum(self.stats.values())
            excel_total, unmatched = update_stats_file(self.stats, today)
            now = datetime.datetime.now().strftime('%H:%M:%S')
            if unmatched or app_total != excel_total:
                missing = ', '.join(f'{k}({v}개)' for k, v in unmatched.items())
                self.lbl_status.config(
                    text=f'{fname}  |  총 {len(self.orders)}건  |  ⚠ 통계 불일치: 앱 {app_total}개 → 엑셀 {excel_total}개  |  누락: {missing}'
                )
            else:
                self.lbl_status.config(
                    text=f'{fname}  |  총 {len(self.orders)}건  |  ✅ 통계 자동저장 완료 — {app_total}개 일치 ({now})'
                )
        except PermissionError:
            self.lbl_status.config(
                text=f'{fname}  |  총 {len(self.orders)}건  |  ⚠ 통계 저장 실패 — 엑셀 파일을 닫아주세요'
            )
        except Exception as e:
            self.lbl_status.config(
                text=f'{fname}  |  총 {len(self.orders)}건  |  ⚠ 통계 저장 오류: {e}'
            )

    def _refresh_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, o in enumerate(self.orders):
            items_str = '  /  '.join(
                f"{k}×{v}" for k, v in
                {it['option']: sum(1 for x in o['items'] if x['option']==it['option'])
                 for it in o['items']}.items()
            )
            price = f"{o['final']:,}원" if o['final'] else '-'
            is_cancel = '취소' in str(o['status'])
            tag = 'cancel' if is_cancel else ('odd' if i % 2 else 'even')
            name_phone = f"{o['reserver']}  {o['phone']}"
            status_label = f'[{o["status"]}]' if is_cancel else '🖨 출력'
            self.tree.insert('', 'end',
                             iid=str(o['order_no']),
                             values=(name_phone, o['use_date'],
                                     items_str, price, o['request'] or '', status_label),
                             tags=(tag,))

    def _refresh_stat(self):
        self.txt_stat.config(state='normal')
        self.txt_stat.delete('1.0', 'end')
        self.txt_stat.insert('end', '  제품별 예약 수량 요약\n', 'header')
        self.txt_stat.insert('end', '  ' + '─' * 38 + '\n', 'dim')
        total = 0
        for prod, cnt in sorted(self.stats.items()):
            self.txt_stat.insert('end', f'  {prod:<28}', 'item')
            self.txt_stat.insert('end', f'{cnt:>4}개\n', 'total')
            total += cnt
        self.txt_stat.insert('end', '  ' + '─' * 38 + '\n', 'dim')
        self.txt_stat.insert('end', f'  {"합  계":<28}{total:>4}개\n', 'total')
        self.txt_stat.config(state='disabled')

    # ── 이벤트 ────────────────────────────────────────────

    def _on_tree_click(self, event):
        if self.tree.identify_region(event.x, event.y) != 'cell':
            return
        col_id = self.tree.identify_column(event.x)
        if col_id != '#6':  # 영수증 열
            return
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return
        order = next((o for o in self.orders if str(o['order_no']) == row_id), None)
        if order:
            self._do_print(order)

    def _do_print(self, order):
        self._show_receipt_preview(order)

    def _show_receipt_preview(self, order):
        win = tk.Toplevel(self)
        win.title(f'영수증 미리보기  —  {order["reserver"]}')
        win.configure(bg=BG_DARK)
        win.geometry('420x580')
        win.resizable(False, False)
        win.grab_set()

        # 제목
        tk.Label(win, text='영수증 미리보기', font=('맑은 고딕', 12, 'bold'),
                 fg=WHITE, bg=BG_DARK).pack(pady=(16, 8))

        # 영수증 종이 느낌 (흰 배경)
        paper = tk.Frame(win, bg='white', padx=20, pady=16,
                         relief='solid', bd=1, width=340)
        paper.pack(padx=30, pady=4, fill='x')
        paper.pack_propagate(False)

        def lbl(text, font_size=11, bold=False, center=False, color='black', pady=2):
            f = ('맑은 고딕', font_size, 'bold') if bold else ('맑은 고딕', font_size)
            anchor = 'center' if center else 'w'
            justify = 'center' if center else 'left'
            tk.Label(paper, text=text, font=f, fg=color, bg='white',
                     anchor=anchor, justify=justify).pack(fill='x', pady=pady)

        def divider(thick=False):
            h = 2 if thick else 1
            c = '#333' if thick else '#ccc'
            tk.Frame(paper, bg=c, height=h).pack(fill='x', pady=4)

        # 영수증 내용
        lbl(f'🥖  {STORE_NAME}', 13, bold=True, center=True, pady=4)
        divider(thick=True)
        lbl(f'예약자:  {order["reserver"]}  {order["phone"]}', 10)
        lbl(f'픽업일시:  {order["use_date"]}', 10)
        divider()

        # 상품 목록 (같은 상품 묶기)
        grouped = OrderedDict()
        for item in order['items']:
            k = item['option']
            if k not in grouped:
                grouped[k] = {'price': item['price'], 'count': 0}
            grouped[k]['count'] += 1

        for opt, v in grouped.items():
            subtotal = v['price'] * v['count']
            row_f = tk.Frame(paper, bg='white')
            row_f.pack(fill='x', pady=1)
            tk.Label(row_f, text=opt, font=('맑은 고딕', 10),
                     fg='black', bg='white', anchor='w').pack(side='left', fill='x', expand=True)
            tk.Label(row_f, text=f'{v["count"]}개', font=('맑은 고딕', 10),
                     fg='#555', bg='white', width=4).pack(side='left')
            tk.Label(row_f, text=f'{subtotal:,}원', font=('맑은 고딕', 10),
                     fg='black', bg='white', anchor='e', width=10).pack(side='right')

        divider(thick=True)
        lbl(f'실결제금액:  {order["final"]:,}원', 12, bold=True)
        lbl(f'결제수단:  {order["payment"]}', 10, color='#555')
        if order['request']:
            divider()
            lbl(f'요청사항:  {order["request"]}', 10, color='#333')
        divider()
        lbl('감사합니다 ♡', 11, center=True, color='#888', pady=6)

        # 버튼 영역
        btn_frame = tk.Frame(win, bg=BG_DARK)
        btn_frame.pack(pady=16)

        def do_print():
            try:
                print_receipt(order)
                self.lbl_status.config(text='브라우저에서 영수증이 열렸습니다 — Ctrl+P로 인쇄하세요')
                win.destroy()
            except Exception as e:
                messagebox.showerror('출력 오류', str(e))

        tk.Button(btn_frame, text='🖨  인쇄하기', command=do_print,
                  bg=ORANGE, fg=WHITE, activebackground='#ff8c5a',
                  relief='flat', font=('맑은 고딕', 11, 'bold'),
                  padx=20, pady=8, cursor='hand2').pack(side='left', padx=8)
        tk.Button(btn_frame, text='✕  닫기', command=win.destroy,
                  bg=BG_ITEM, fg=WHITE, activebackground='#3a3a6e',
                  relief='flat', font=('맑은 고딕', 11),
                  padx=20, pady=8, cursor='hand2').pack(side='left', padx=8)

    def _print_all_orders(self):
        if not self.orders:
            messagebox.showwarning('경고', '예약 데이터가 없습니다.\n먼저 예약 파일을 불러와 주세요.')
            return
        html = make_order_list_html(self.orders)
        tmp = tempfile.NamedTemporaryFile(
            mode='w', suffix='.html', encoding='utf-8', delete=False,
            dir=tempfile.gettempdir()
        )
        tmp.write(html)
        tmp.close()
        open_in_browser(tmp.name)
        self.lbl_status.config(text='브라우저에서 예약 목록이 열렸습니다 — Ctrl+P로 인쇄하세요')

    def _update_stats(self):
        if not self.stats:
            messagebox.showwarning('경고', '예약 데이터가 없습니다.\n먼저 예약 파일을 불러와 주세요.')
            return
        if not os.path.exists(STATS_FILE):
            messagebox.showerror('오류', f'통계 파일을 찾을 수 없습니다:\n{STATS_FILE}')
            return
        try:
            today = datetime.date.today().strftime('%Y년 %m월 %d일')
            app_total = sum(self.stats.values())
            excel_total, unmatched = update_stats_file(self.stats, today)
            now = datetime.datetime.now().strftime('%H:%M:%S')

            # 갯수 일치 여부 확인
            ok = (app_total == excel_total and not unmatched)
            match_line = (
                f'  ✅ 앱 통계 {app_total}개  =  엑셀 저장 {excel_total}개  (일치)'
                if ok else
                f'  ⚠ 앱 통계 {app_total}개  ≠  엑셀 저장 {excel_total}개  (불일치)'
            )
            unmatched_line = ''
            if unmatched:
                items = '\n'.join(f'    · {k} ({v}개)' for k, v in unmatched.items())
                unmatched_line = f'\n\n⚠ 통계 파일에 없는 예약 제품 (엑셀 A열 추가 필요):\n{items}'

            self.lbl_status.config(
                text=f'{"✅" if ok else "⚠"} 통계 저장 완료 ({now})  |  앱 {app_total}개 / 엑셀 {excel_total}개'
            )
            ans = messagebox.askyesno(
                '통계 파일 업데이트 완료',
                f'[형만님] 시트 B열에 저장됐습니다.\n\n'
                f'  저장 시각: {today}  {now}\n'
                f'  상품 종류: {len(self.stats)}가지\n'
                f'{match_line}'
                f'{unmatched_line}\n\n'
                f'지금 엑셀 파일을 열어볼까요?'
            )
            if ans:
                os.startfile(STATS_FILE)
        except PermissionError:
            messagebox.showerror('저장 실패',
                '통계 파일이 현재 열려 있습니다.\n'
                '엑셀을 닫고 다시 시도해주세요.')
        except Exception as e:
            messagebox.showerror('오류', str(e))


if __name__ == '__main__':
    app = App()
    app.mainloop()
